import os
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from flask import current_app  # Import current_app here

def allowed_file(filename):
    """Check if the uploaded file is allowed based on its extension."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'xlsx'}

def format_output_file(output_file_path):
    """Format the output Excel file with headers and auto-width columns."""
    wb = openpyxl.load_workbook(output_file_path)
    
    for sheet in wb.worksheets:
        ws = wb[sheet.title]
        
        # Set the header format
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        
        for col in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col)
            ws[f'{col_letter}1'].font = header_font
            ws[f'{col_letter}1'].fill = header_fill
            
            # Auto-width for columns
            max_length = 0
            for row in range(1, ws.max_row + 1):
                cell_value = ws[f'{col_letter}{row}'].value
                if cell_value is not None:
                    cell_length = len(str(cell_value))
                    if cell_length > max_length:
                        max_length = cell_length
            
            adjusted_width = max_length + 2
            ws.column_dimensions[col_letter].width = adjusted_width

    wb.save(output_file_path)

def calculate_total_qty(file_path):
    """Calculate the total quantity needed for each branch based on uploaded data."""
    try:
        # Load the workbook uploaded by the user
        user_wb = openpyxl.load_workbook(file_path, data_only=True)
        
        # Load the reference workbook with zero quantities
        ref_wb = openpyxl.load_workbook('static/zeros_data.xlsx', data_only=True)
        
        output_data = {'total_branches': [], 'missing_items': []}
        branch_data = {}

        for sheet_name in ref_wb.sheetnames:
            ref_sheet = ref_wb[sheet_name]

            # Initialize lists to store reference data
            ref_ids = []
            ref_names = []
            zero_qty = []
            ref_source = []

            # Extract data from reference sheet starting from the second row
            for row in ref_sheet.iter_rows(min_row=2, min_col=1, max_col=4):
                ref_ids.append(row[0].value)
                ref_names.append(row[1].value)
                zero_qty.append(row[2].value)
                ref_source.append(row[3].value)  # Assume the source column is the fourth column

            # Create DataFrame from extracted reference data
            ref_df = pd.DataFrame({'ID': ref_ids, 'Name': ref_names, 'Zero_Qty': zero_qty, 'Source': ref_source})
            ref_df['Zero_Qty'] = pd.to_numeric(ref_df['Zero_Qty'], errors='coerce')

            # Check if user workbook has the current sheet
            if sheet_name in user_wb.sheetnames:
                user_sheet = user_wb[sheet_name]
                
                # Initialize lists to store data
                ids = []
                components = []
                total_used = []

                # Extract data from user sheet starting from the second row
                for row in user_sheet.iter_rows(min_row=2, min_col=1, max_col=3):
                    ids.append(row[0].value)
                    components.append(row[1].value)
                    total_used.append(row[2].value)

                # Create DataFrame from extracted data
                user_df = pd.DataFrame({'ID': ids, 'Component': components, 'Total Used': total_used})
                user_df['Total Used'] = pd.to_numeric(user_df['Total Used'], errors='coerce')
            else:
                # If user sheet doesn't exist, create an empty DataFrame
                user_df = pd.DataFrame(columns=['ID', 'Component', 'Total Used'])

            # Find missing IDs in user input
            missing_in_user_input = ref_df[~ref_df['ID'].isin(user_df['ID'])]

            # Add missing items to user DataFrame with Total Used as 0
            missing_in_user_input['Total Used'] = 0
            missing_in_user_input = missing_in_user_input.rename(columns={'Name': 'Component'})
            user_df = pd.concat([user_df, missing_in_user_input[['ID', 'Component', 'Total Used']]], ignore_index=True)

            # Merge dataframes on 'ID' column
            merged_df = pd.merge(ref_df, user_df, on='ID', how='left')

            # Identify missing items not found in reference (zeros_data.xlsx)
            missing_items_df = user_df[~user_df['ID'].isin(ref_df['ID'])]
            if not missing_items_df.empty:
                missing_items_df['Qty_Needed'] = -missing_items_df['Total Used']
                missing_items_df['Branch'] = sheet_name
                output_data['missing_items'].append(missing_items_df[['ID', 'Component', 'Qty_Needed', 'Branch']])

            # Perform calculations only for items found in reference data
            merged_df = merged_df.dropna(subset=['Zero_Qty'])
            merged_df['Qty_Needed'] = merged_df['Zero_Qty'] - merged_df['Total Used']

            # Store data for individual branches
            branch_data[sheet_name] = merged_df[['ID', 'Name', 'Qty_Needed', 'Source']]
            
            # Append data to total_branches
            output_data['total_branches'].append(merged_df[['ID', 'Name', 'Qty_Needed', 'Source']])

        # Combine total_branches data into a single DataFrame
        total_branches_df = pd.concat(output_data['total_branches']).groupby(['ID', 'Name', 'Source'], as_index=False).sum()

        # Split into warehouse and kitchen based on 'Source' column
        warehouse_df = total_branches_df[total_branches_df['Source'] == 'warehouse']
        kitchen_df = total_branches_df[total_branches_df['Source'] == 'kitchen']

        # Combine missing_items data into a single DataFrame
        if output_data['missing_items']:
            missing_items_df = pd.concat(output_data['missing_items']).drop_duplicates()
        else:
            missing_items_df = pd.DataFrame(columns=['ID', 'Component', 'Qty_Needed', 'Branch'])

        # Create Excel workbooks for warehouse and kitchen outputs
        warehouse_file_path = os.path.join(current_app.config['UPLOAD_FOLDER'], 'warehouse_total_qty.xlsx')
        kitchen_file_path = os.path.join(current_app.config['UPLOAD_FOLDER'], 'kitchen_total_qty.xlsx')

        with pd.ExcelWriter(warehouse_file_path) as writer:
            warehouse_df.to_excel(writer, sheet_name='total_branches', index=False)
            missing_items_df.to_excel(writer, sheet_name='missing_items', index=False)
            for branch, df in branch_data.items():
                df[df['Source'] == 'warehouse'].to_excel(writer, sheet_name=branch, index=False)
        
        with pd.ExcelWriter(kitchen_file_path) as writer:
            kitchen_df.to_excel(writer, sheet_name='total_branches', index=False)
            missing_items_df.to_excel(writer, sheet_name='missing_items', index=False)
            for branch, df in branch_data.items():
                df[df['Source'] == 'kitchen'].to_excel(writer, sheet_name=branch, index=False)

        # Format the output files
        format_output_file(warehouse_file_path)
        format_output_file(kitchen_file_path)

        return warehouse_file_path, kitchen_file_path

    except Exception as e:
        return f"An error occurred while processing: {e}"
