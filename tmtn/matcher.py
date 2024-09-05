import os
import openpyxl
from flask import current_app, url_for
from flask_socketio import emit
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
import pandas as pd
from openpyxl.styles import PatternFill, Alignment, Font
from .helper import allowed_file


def daily_matcher(file_path):
    try:
        # Load the workbook uploaded by the user
        wb = openpyxl.load_workbook(file_path, data_only=True)

        # Read all the relevant sheets into DataFrames
        sheets = ["start_A", "add_materials", "daily_consumed", "canceled", "damaged", "end_C"]
        data = {sheet: pd.DataFrame(wb[sheet].iter_rows(min_row=2, values_only=True), columns=["ID", "NAME", "QTY"]) for sheet in sheets}

        # Create an empty DataFrame for total quantities and individual sheet quantities
        total_qty_df = pd.DataFrame(columns=["ID", "NAME", "start_A QTY", "add_materials QTY", "canceled QTY", "daily_consumed QTY", "damaged QTY", "end_C QTY", "Total QTY"])

        # Get unique IDs from all sheets
        unique_ids = pd.concat([data[sheet]['ID'] for sheet in sheets]).unique()

        for id in unique_ids:
            name = None
            total_qty = 0

            # Initialize quantities for each sheet
            start_A_qty = 0
            add_materials_qty = 0
            canceled_qty = 0
            daily_consumed_qty = 0
            damaged_qty = 0
            end_C_qty = 0

            for sheet in sheets:
                sheet_data = data[sheet][data[sheet]['ID'] == id]
                if not sheet_data.empty:
                    name = sheet_data.iloc[0]['NAME']
                    qty = sheet_data.iloc[0]['QTY']
                    
                    if sheet == "start_A":
                        start_A_qty += qty
                        total_qty += qty
                    elif sheet == "add_materials":
                        add_materials_qty += qty
                        total_qty += qty
                    elif sheet == "canceled":
                        canceled_qty += qty
                        total_qty += qty
                    elif sheet == "daily_consumed":
                        daily_consumed_qty += qty
                        total_qty -= qty
                    elif sheet == "damaged":
                        damaged_qty += qty
                        total_qty -= qty
                    elif sheet == "end_C":
                        end_C_qty += qty
                        total_qty -= qty

            # Append row data to total_qty_df
            total_qty_df = pd.concat([total_qty_df, pd.DataFrame([{
                "ID": id, 
                "NAME": name, 
                "start_A QTY": start_A_qty,
                "add_materials QTY": add_materials_qty,
                "canceled QTY": canceled_qty,
                "daily_consumed QTY": daily_consumed_qty,
                "damaged QTY": damaged_qty,
                "end_C QTY": end_C_qty,
                "Total QTY": total_qty
            }])], ignore_index=True)

        # Save the result to a new sheet in the Excel file
        output_file_path = os.path.join(current_app.config['UPLOAD_FOLDER'], 'final_matcher.xlsx')
        with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
            total_qty_df.to_excel(writer, sheet_name='total_qty', index=False)

        # Load the workbook for formatting
        wb = openpyxl.load_workbook(output_file_path)
        ws = wb['total_qty']

        def format_sheet(sheet):
            # Apply color fills
            green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
            red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

            # Column indices (1-based index)
            start_A_col = 3
            add_materials_col = 4
            canceled_col = 5
            daily_consumed_col = 6
            damaged_col = 7
            end_C_col = 8

            for col in range(1, sheet.max_column + 1):
                max_length = 0
                column = sheet.cell(row=1, column=col).column_letter
                for row in range(1, sheet.max_row + 1):
                    cell = sheet.cell(row=row, column=col)
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                    
                    if col == start_A_col or col == add_materials_col or col == canceled_col:
                        if row > 1:  # Skip the header row
                            cell.fill = green_fill
                    elif col == daily_consumed_col or col == damaged_col or col == end_C_col:
                        if row > 1:  # Skip the header row
                            cell.fill = red_fill

                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column].width = adjusted_width

            # Apply modern header formatting
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            header_alignment = Alignment(horizontal="center", vertical="center")
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
        
        format_sheet(ws)
        wb.save(output_file_path)

        return output_file_path

    except Exception as e:
        return f"An error occurred while processing: {e}"
    
def register_matcher_routes(app, socketio):
    @socketio.on('file_uploaded_matcher')  # Unique event name for matcher
    def handle_file_uploaded_matcher(data):
        try:
            filename = data['filename']
            file_content = data['data']  # Assume this is already in bytes

            # Convert the file data back to a file-like object (for processing)
            file_path2 = os.path.join(current_app.config['UPLOAD_FOLDER'], secure_filename(filename))

            # Ensure the upload folder exists before writing
            os.makedirs(os.path.dirname(file_path2), exist_ok=True)

            # Save the uploaded file data to disk (using binary mode)
            with open(file_path2, 'wb') as f:
                f.write(file_content)

            socketio.emit('processing_status_matcher', {'status': 'processing', 'message': 'جارٍ معالجة الملف...'})

            processed_file_path = daily_matcher(file_path2)
            if processed_file_path.startswith('An error occurred'):
                socketio.emit('processing_status_matcher', {'status': 'error', 'message': 'حدث خطأ أثناء معالجة الملف.'})
            else:
                socketio.emit('processing_status_matcher', {
                    'status': 'success',
                    'message': 'تمت معالجة الملف بنجاح!',
                    'file_url': url_for('download_file', filename='final_matcher.xlsx', _external=True)
                })
        except Exception as e:
            print(f"Error processing file_uploaded_matcher: {e}")  # Log the error for debugging
            socketio.emit('processing_status_matcher', {'status': 'error', 'message': f'حدث خطأ: {str(e)}'})