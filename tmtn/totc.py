import os
from flask import current_app, url_for
from flask_socketio import emit
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
import pandas as pd
from openpyxl.styles import PatternFill, Alignment, Font
from .helper import allowed_file



def extract_and_calculate_usage(file_path1):
    try:
        # Ensure static folder exists
        static_file_path = os.path.join(current_app.config['STATIC_FOLDER'], 'recipe_data.xlsx')
        if not os.path.exists(static_file_path):
            raise FileNotFoundError(f"Static file {static_file_path} not found.")

        # Load the daily sales workbook (uploaded by user)
        wb1 = load_workbook(file_path1, data_only=True)
        ws1 = wb1[wb1.sheetnames[0]]
        data1 = ws1.iter_rows(min_row=2, values_only=True)  # Start from the second row, assuming the first row is the header
        sold_drinks_data = pd.DataFrame(data1, columns=["ID", "Drink", "Qty", "Shift"])

        # Load the static recipe workbook
        wb2 = load_workbook(static_file_path, data_only=True)
        ws2 = wb2[wb2.sheetnames[0]]
        data2 = ws2.iter_rows(min_row=2, values_only=True)  # Start from the second row, assuming the first row is the header
        recipe_data = pd.DataFrame(data2, columns=["ID", "NAME", "Package", "QTY"])

        # Normalize names for consistency
        recipe_data['NAME'] = recipe_data['NAME'].str.strip().str.lower()
        sold_drinks_data['Drink'] = sold_drinks_data['Drink'].str.strip().str.lower()

        # Normalize shift category to 'A', 'B', or 'C'
        sold_drinks_data['ShiftCategory'] = sold_drinks_data['Shift'].str[0]

        # Create a dictionary to map each drink to its components
        components_dict = {}
        current_drink = None

        for index, row in recipe_data.iterrows():
            if row['Package'] == 'drink':
                current_drink = row['NAME']
                components_dict[current_drink] = []
            else:
                if current_drink:
                    components_dict[current_drink].append({'ID': row['ID'], 'NAME': row['NAME'], 'QTY': row['QTY']})

        # Initialize dictionaries for total and shift-specific components
        total_components = {}
        shift_components = {'A': {}, 'B': {}, 'C': {}}
        missing_items = []

        # Calculate total components used for each shift category
        for index, row in sold_drinks_data.iterrows():
            drink_name = row['Drink']
            qty_sold = row['Qty']
            shift_category = row['ShiftCategory']

            if drink_name in components_dict:
                for component in components_dict[drink_name]:
                    component_id = component['ID']
                    component_name = component['NAME']
                    component_qty = component['QTY'] * qty_sold

                    # Update total components
                    if component_id in total_components:
                        total_components[component_id]['Total Used'] += component_qty
                    else:
                        total_components[component_id] = {'Component': component_name, 'Total Used': component_qty}

                    # Update shift-specific components
                    if component_id in shift_components[shift_category]:
                        shift_components[shift_category][component_id]['Total Used'] += component_qty
                    else:
                        shift_components[shift_category][component_id] = {'Component': component_name, 'Total Used': component_qty}
            else:
                # Add missing items to the list
                missing_items.append({'ID': row['ID'], 'Drink': row['Drink'], 'Qty': row['Qty'], 'Shift': row['Shift']})

        # Convert total components and shift-specific components to DataFrames
        total_components_df = pd.DataFrame([(id, data['Component'], data['Total Used']) for id, data in total_components.items()], 
                                           columns=['ID', 'Component', 'Total Used'])
        shift_A_df = pd.DataFrame([(id, data['Component'], data['Total Used']) for id, data in shift_components['A'].items()], 
                                   columns=['ID', 'Component', 'Total Used'])
        shift_B_df = pd.DataFrame([(id, data['Component'], data['Total Used']) for id, data in shift_components['B'].items()], 
                                   columns=['ID', 'Component', 'Total Used'])
        shift_C_df = pd.DataFrame([(id, data['Component'], data['Total Used']) for id, data in shift_components['C'].items()], 
                                   columns=['ID', 'Component', 'Total Used'])

        # Create a new Excel writer to save the output
        output_file_path = os.path.join(current_app.config['UPLOAD_FOLDER'], 'final_content_usage.xlsx')
        with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
            # Save the dataframes to separate sheets
            total_components_df.to_excel(writer, sheet_name='TotalUsage', index=False)
            shift_A_df.to_excel(writer, sheet_name='A', index=False)
            shift_B_df.to_excel(writer, sheet_name='B', index=False)
            shift_C_df.to_excel(writer, sheet_name='C', index=False)
            
            # Conditionally add the missing items sheet if there are missing items
            if missing_items:
                missing_items_df = pd.DataFrame(missing_items, columns=['ID', 'Drink', 'Qty', 'Shift'])
                missing_items_df.to_excel(writer, sheet_name='MissingItems', index=False)

        # Apply formatting to the Excel file
        format_excel_file(output_file_path)

        # Return the path to the processed file
        return output_file_path

    except Exception as e:
        print(f"Error in extract_and_calculate_usage: {e}")  # Log the error for debugging
        return f"An error occurred while processing: {e}"

def format_excel_file(file_path):
    """Apply formatting to an Excel file."""
    # Load the workbook
    wb = load_workbook(file_path)
    
    # Define the header style
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for sheet in wb.sheetnames:
        ws = wb[sheet]

        # Apply the header style
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

    # Save the formatted workbook
    wb.save(file_path)


def register_totc_routes(app, socketio):
    @socketio.on('file_uploaded_totc')  # Unique event name for TOTC
    def handle_file_uploaded_totc(data):
        try:
            filename = data['filename']
            file_content = data['data']  # Assume this is already in bytes

            # Convert the file data back to a file-like object (for processing)
            file_path1 = os.path.join(current_app.config['UPLOAD_FOLDER'], secure_filename(filename))

            # Ensure the upload folder exists before writing
            os.makedirs(os.path.dirname(file_path1), exist_ok=True)

            # Save the uploaded file data to disk (using binary mode)
            with open(file_path1, 'wb') as f:
                f.write(file_content)

            socketio.emit('processing_status_totc', {'status': 'processing', 'message': 'جارٍ معالجة الملف...'})

            processed_file_path = extract_and_calculate_usage(file_path1)
            if processed_file_path.startswith('An error occurred'):
                socketio.emit('processing_status_totc', {'status': 'error', 'message': 'حدث خطأ أثناء معالجة الملف.'})
            else:
                socketio.emit('processing_status_totc', {
                    'status': 'success',
                    'message': 'تمت معالجة الملف بنجاح!',
                    'file_url': url_for('download_file', filename='final_content_usage.xlsx', _external=True)
                })
        except Exception as e:
            print(f"Error processing file_uploaded_totc: {e}")  # Log the error for debugging
            socketio.emit('processing_status_totc', {'status': 'error', 'message': f'حدث خطأ: {str(e)}'})